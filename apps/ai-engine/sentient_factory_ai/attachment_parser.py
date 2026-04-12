from __future__ import annotations

import csv
import io
import shutil
import subprocess
import tempfile
from pathlib import Path
from zipfile import ZipFile

from fastapi import UploadFile

from .models import ChatAttachment

MAX_ATTACHMENT_BYTES = 15 * 1024 * 1024
CONTENT_LIMIT = 12_000


def _clamp_text(value: str, limit: int = CONTENT_LIMIT) -> str:
    normalized = value.replace("\x00", " ")
    normalized = " ".join(normalized.split()).strip()
    if len(normalized) <= limit:
        return normalized
    return f"{normalized[:limit]}…"


def _strip_xml(value: str) -> str:
    import re

    cleaned = re.sub(r"<[^>]+>", " ", value)
    return (
        cleaned.replace("&amp;", "&")
        .replace("&lt;", "<")
        .replace("&gt;", ">")
        .replace("&quot;", '"')
        .replace("&apos;", "'")
    )


def _build_preview(value: str) -> str:
    preview = _clamp_text(value, 240)
    return preview or "Tidak ada teks yang berhasil diekstrak."


def _get_extension(filename: str) -> str:
    parts = filename.lower().split(".")
    return parts[-1] if len(parts) > 1 else ""


def _make_attachment(
    *,
    upload: UploadFile,
    size_bytes: int,
    extension: str,
    status: str,
    content: str,
    warning: str | None = None,
    metadata: dict[str, object] | None = None,
) -> ChatAttachment:
    normalized_content = _clamp_text(content)
    return ChatAttachment(
        name=upload.filename or "attachment",
        media_type=upload.content_type,
        size_bytes=size_bytes,
        extension=extension,
        extraction_status=status,  # type: ignore[arg-type]
        content=normalized_content,
        preview=_build_preview(normalized_content),
        warning=warning,
        metadata={key: value for key, value in (metadata or {}).items()},
    )


def _extract_text_file(raw_bytes: bytes, encoding: str = "utf-8") -> str:
    return raw_bytes.decode(encoding, errors="ignore")


def _extract_csv(raw_bytes: bytes) -> tuple[str, dict[str, object]]:
    decoded = raw_bytes.decode("utf-8", errors="ignore")
    reader = csv.reader(io.StringIO(decoded))
    rows = []
    row_count = 0
    for row in reader:
        values = [cell.strip() for cell in row if cell.strip()]
        if values:
            row_count += 1
            rows.append(" | ".join(values))
    return "\n".join(rows), {"row_count": row_count}


def _extract_docx(raw_bytes: bytes) -> tuple[str, dict[str, object]]:
    texts: list[str] = []
    with ZipFile(io.BytesIO(raw_bytes)) as archive:
        candidates = ["word/document.xml"] + sorted(
            name
            for name in archive.namelist()
            if name.startswith("word/header") or name.startswith("word/footer")
        )
        for name in candidates:
            try:
                xml = archive.read(name).decode("utf-8", errors="ignore")
            except KeyError:
                continue
            stripped = _strip_xml(xml)
            if stripped.strip():
                texts.append(stripped)
    return "\n".join(texts), {"section_count": len(texts)}


def _extract_xlsx(raw_bytes: bytes) -> tuple[str, dict[str, object]]:
    try:
        from openpyxl import load_workbook
    except Exception as error:  # pragma: no cover
        raise RuntimeError("openpyxl belum terpasang di ai-engine.") from error

    workbook = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    lines: list[str] = []
    row_count = 0
    for worksheet in workbook.worksheets:
        lines.append(f"Sheet: {worksheet.title}")
        for row in worksheet.iter_rows(values_only=True):
            values = [str(value).strip() for value in row if value is not None and str(value).strip()]
            if values:
                row_count += 1
                lines.append(" | ".join(values))
    return "\n".join(lines), {"sheet_count": len(workbook.worksheets), "row_count": row_count}


def _run_tesseract(image_path: Path) -> str:
    if shutil.which("tesseract") is None:
        raise RuntimeError("Binary tesseract tidak tersedia di server.")

    command = ["tesseract", str(image_path), "stdout", "-l", "ind+eng"]
    completed = subprocess.run(
        command,
        check=False,
        capture_output=True,
        text=True,
    )
    if completed.returncode != 0:
        stderr = completed.stderr.strip() or "Tesseract gagal mengekstrak teks."
        raise RuntimeError(stderr)
    return completed.stdout


def _ocr_image(raw_bytes: bytes, extension: str) -> tuple[str, dict[str, object]]:
    try:
        from PIL import Image
    except Exception as error:  # pragma: no cover
        raise RuntimeError("Pillow belum terpasang di ai-engine.") from error

    with tempfile.TemporaryDirectory(prefix="sf-ocr-img-") as tmp_dir:
        image_path = Path(tmp_dir) / f"input.{extension or 'png'}"
        image_path.write_bytes(raw_bytes)
        with Image.open(image_path) as image:
            width, height = image.size
        text = _run_tesseract(image_path)
    return text, {"width": width, "height": height, "ocr_engine": "tesseract"}


def _extract_pdf_text(raw_bytes: bytes) -> str:
    try:
        from pypdf import PdfReader
    except Exception as error:  # pragma: no cover
        raise RuntimeError("pypdf belum terpasang di ai-engine.") from error

    reader = PdfReader(io.BytesIO(raw_bytes))
    texts: list[str] = []
    for page in reader.pages:
        page_text = page.extract_text() or ""
        if page_text.strip():
            texts.append(page_text)
    return "\n".join(texts)


def _ocr_pdf(raw_bytes: bytes) -> tuple[str, dict[str, object]]:
    if shutil.which("pdftoppm") is None:
        raise RuntimeError("Binary pdftoppm tidak tersedia di server.")

    with tempfile.TemporaryDirectory(prefix="sf-ocr-pdf-") as tmp_dir:
        pdf_path = Path(tmp_dir) / "input.pdf"
        pdf_path.write_bytes(raw_bytes)
        prefix = Path(tmp_dir) / "page"
        command = ["pdftoppm", "-png", str(pdf_path), str(prefix)]
        completed = subprocess.run(
            command,
            check=False,
            capture_output=True,
            text=True,
        )
        if completed.returncode != 0:
            stderr = completed.stderr.strip() or "pdftoppm gagal merender PDF."
            raise RuntimeError(stderr)

        images = sorted(Path(tmp_dir).glob("page-*.png"))
        texts: list[str] = []
        for image_path in images:
            text = _run_tesseract(image_path)
            if text.strip():
                texts.append(text)
    return "\n".join(texts), {"page_count": len(images), "ocr_engine": "tesseract"}


async def parse_upload_attachment(upload: UploadFile) -> ChatAttachment:
    raw_bytes = await upload.read()
    size_bytes = len(raw_bytes)
    extension = _get_extension(upload.filename or "")

    if size_bytes > MAX_ATTACHMENT_BYTES:
        return _make_attachment(
            upload=upload,
            size_bytes=size_bytes,
            extension=extension,
            status="failed",
            content="",
            warning="File terlalu besar. Maksimum 15 MB per file.",
        )

    try:
        if (upload.content_type or "").startswith("text/") or extension in {"txt", "md", "json", "xml", "html"}:
            text = _extract_text_file(raw_bytes)
            return _make_attachment(
                upload=upload,
                size_bytes=size_bytes,
                extension=extension,
                status="ready",
                content=text,
            )

        if extension == "csv":
            text, metadata = _extract_csv(raw_bytes)
            return _make_attachment(
                upload=upload,
                size_bytes=size_bytes,
                extension=extension,
                status="ready" if text.strip() else "metadata-only",
                content=text,
                metadata=metadata,
            )

        if extension in {"xlsx", "xlsm"} or upload.content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            text, metadata = _extract_xlsx(raw_bytes)
            return _make_attachment(
                upload=upload,
                size_bytes=size_bytes,
                extension=extension,
                status="ready" if text.strip() else "metadata-only",
                content=text,
                metadata=metadata,
                warning=None if text.strip() else "Workbook terbaca tetapi tidak menghasilkan teks yang berarti.",
            )

        if extension == "docx" or upload.content_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            text, metadata = _extract_docx(raw_bytes)
            return _make_attachment(
                upload=upload,
                size_bytes=size_bytes,
                extension=extension,
                status="ready" if text.strip() else "metadata-only",
                content=text,
                metadata=metadata,
                warning=None if text.strip() else "Dokumen Word terbaca tetapi tidak menghasilkan teks yang berarti.",
            )

        if extension == "pdf" or upload.content_type == "application/pdf":
            text = _extract_pdf_text(raw_bytes)
            if text.strip():
                return _make_attachment(
                    upload=upload,
                    size_bytes=size_bytes,
                    extension=extension,
                    status="ready",
                    content=text,
                    warning="PDF dibaca dari layer teks native.",
                    metadata={"extraction_mode": "native-text"},
                )

            ocr_text, metadata = _ocr_pdf(raw_bytes)
            return _make_attachment(
                upload=upload,
                size_bytes=size_bytes,
                extension=extension,
                status="ready" if ocr_text.strip() else "metadata-only",
                content=ocr_text,
                warning="PDF di-OCR dengan Tesseract karena layer teks native tidak ditemukan.",
                metadata={"extraction_mode": "ocr", **metadata},
            )

        if (upload.content_type or "").startswith("image/") or extension in {"png", "jpg", "jpeg", "webp", "gif", "bmp"}:
            text, metadata = _ocr_image(raw_bytes, extension or "png")
            return _make_attachment(
                upload=upload,
                size_bytes=size_bytes,
                extension=extension,
                status="ready" if text.strip() else "metadata-only",
                content=text,
                warning="Image diproses dengan OCR Tesseract.",
                metadata=metadata,
            )

        if extension == "doc":
            return _make_attachment(
                upload=upload,
                size_bytes=size_bytes,
                extension=extension,
                status="metadata-only",
                content=f"Dokumen Word legacy {upload.filename or 'attachment'} diterima, tetapi parsing .doc belum tersedia.",
                warning="Format .doc belum didukung. Gunakan .docx agar isi dokumen bisa diekstrak.",
            )

        if extension == "xls":
            return _make_attachment(
                upload=upload,
                size_bytes=size_bytes,
                extension=extension,
                status="metadata-only",
                content=f"Workbook Excel legacy {upload.filename or 'attachment'} diterima, tetapi parsing .xls belum tersedia.",
                warning="Format .xls belum didukung. Gunakan .xlsx agar isi workbook bisa diekstrak.",
            )

        return _make_attachment(
            upload=upload,
            size_bytes=size_bytes,
            extension=extension,
            status="metadata-only",
            content=f"File {upload.filename or 'attachment'} diterima, tetapi tipe file ini belum punya parser khusus.",
            warning="Tipe file belum didukung untuk ekstraksi konten. Hanya metadata file yang diteruskan.",
        )
    except Exception as error:
        return _make_attachment(
            upload=upload,
            size_bytes=size_bytes,
            extension=extension,
            status="failed",
            content="",
            warning=str(error),
        )
